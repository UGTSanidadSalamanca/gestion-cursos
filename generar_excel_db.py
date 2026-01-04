#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Excel con la estructura completa de la base de datos
Sistema de Gestión Integral de Cursos
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_database_excel():
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def apply_header_style(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
    
    def auto_adjust_columns(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # HOJA 1: RESUMEN
    ws1 = wb.active
    ws1.title = "Resumen de Tablas"
    ws1.append(["Tabla", "Nombre en BD", "Descripción", "Campos Principales", "Relaciones"])
    apply_header_style(ws1)
    
    resumen_data = [
        ["User", "users", "Usuarios del sistema", "id | email | name | role", "→ Contact | → Notification"],
        ["Student", "students", "Alumnos afiliados y no afiliados", "id | name | email | dni | status", "→ Enrollment | → Payment | → Contact"],
        ["Teacher", "teachers", "Docentes e instructores", "id | name | email | specialty | status", "→ Course | → Contact"],
        ["Provider", "providers", "Proveedores de materiales", "id | name | taxId | category | status", "→ Material | → Contact"],
        ["Course", "courses", "Cursos ofrecidos", "id | title | code | level | price", "← Teacher | → Enrollment | → Schedule"],
        ["Enrollment", "enrollments", "Matrículas de estudiantes", "id | studentId | courseId | status", "← Student | ← Course"],
        ["Payment", "payments", "Pagos y transacciones", "id | amount | paymentMethod | status", "← Student | ← Course"],
        ["Material", "materials", "Materiales educativos", "id | name | type | quantity | price", "← Provider | → CourseMaterial"],
        ["CourseMaterial", "course_materials", "Relación Cursos-Materiales", "courseId | materialId | quantity", "← Course | ← Material"],
        ["Schedule", "schedules", "Horarios de clases", "courseId | dayOfWeek | startTime", "← Course"],
        ["Contact", "contacts", "Contactos polimórficos", "name | email | phone | category", "← Student | ← Teacher | ← Provider"],
        ["Software", "software", "Software y plataformas", "name | type | license | expiryDate", "Sin relaciones"],
        ["Notification", "notifications", "Notificaciones del sistema", "title | message | type | priority", "← User"],
        ["NotificationPreference", "notification_preferences", "Preferencias de notificación", "userId | category | enabled", "← User"],
    ]
    
    for row in resumen_data:
        ws1.append(row)
    
    auto_adjust_columns(ws1)
    
    # HOJA 2: USER
    ws2 = wb.create_sheet("User")
    ws2.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws2)
    
    user_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["email", "String", "Sí", "Sí", "-", "Correo electrónico", "-"],
        ["name", "String", "No", "No", "-", "Nombre completo", "-"],
        ["role", "UserRole", "Sí", "No", "ADMIN", "Rol (ADMIN|TEACHER|STAFF)", "-"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["contacts", "Contact[]", "No", "No", "-", "Contactos asociados", "1:N → Contact"],
        ["notifications", "Notification[]", "No", "No", "-", "Notificaciones", "1:N → Notification"],
        ["notificationPreferences", "NotificationPreference[]", "No", "No", "-", "Preferencias", "1:N → NotificationPreference"],
    ]
    
    for row in user_data:
        ws2.append(row)
    auto_adjust_columns(ws2)
    
    # HOJA 3: STUDENT
    ws3 = wb.create_sheet("Student")
    ws3.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws3)
    
    student_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["name", "String", "Sí", "No", "-", "Nombre completo", "-"],
        ["email", "String", "Sí", "Sí", "-", "Correo electrónico", "-"],
        ["phone", "String", "No", "No", "-", "Teléfono", "-"],
        ["address", "String", "No", "No", "-", "Dirección", "-"],
        ["dni", "String", "No", "Sí", "-", "DNI/Documento", "-"],
        ["birthDate", "DateTime", "No", "No", "-", "Fecha de nacimiento", "-"],
        ["isAffiliated", "Boolean", "Sí", "No", "false", "Es afiliado", "-"],
        ["affiliateNumber", "String", "No", "Sí", "-", "Número de afiliado", "-"],
        ["emergencyContact", "String", "No", "No", "-", "Contacto emergencia", "-"],
        ["emergencyPhone", "String", "No", "No", "-", "Teléfono emergencia", "-"],
        ["medicalInfo", "String", "No", "No", "-", "Info médica", "-"],
        ["status", "StudentStatus", "Sí", "No", "ACTIVE", "Estado (ACTIVE|INACTIVE|SUSPENDED|GRADUATED)", "-"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["enrollments", "Enrollment[]", "No", "No", "-", "Matrículas", "1:N → Enrollment"],
        ["payments", "Payment[]", "No", "No", "-", "Pagos", "1:N → Payment"],
        ["contacts", "Contact[]", "No", "No", "-", "Contactos", "1:N → Contact"],
    ]
    
    for row in student_data:
        ws3.append(row)
    auto_adjust_columns(ws3)
    
    # HOJA 4: TEACHER
    ws4 = wb.create_sheet("Teacher")
    ws4.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws4)
    
    teacher_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["name", "String", "Sí", "No", "-", "Nombre completo", "-"],
        ["email", "String", "Sí", "Sí", "-", "Correo electrónico", "-"],
        ["phone", "String", "No", "No", "-", "Teléfono", "-"],
        ["address", "String", "No", "No", "-", "Dirección", "-"],
        ["dni", "String", "No", "Sí", "-", "DNI/Documento", "-"],
        ["specialty", "String", "No", "No", "-", "Especialidad", "-"],
        ["experience", "String", "No", "No", "-", "Experiencia", "-"],
        ["cv", "String", "No", "No", "-", "URL del CV", "-"],
        ["contractType", "ContractType", "Sí", "No", "FREELANCE", "Tipo (FREELANCE|PART_TIME|FULL_TIME|HOURLY)", "-"],
        ["hourlyRate", "Float", "No", "No", "-", "Tarifa por hora", "-"],
        ["status", "TeacherStatus", "Sí", "No", "ACTIVE", "Estado (ACTIVE|INACTIVE|ON_LEAVE)", "-"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["courses", "Course[]", "No", "No", "-", "Cursos impartidos", "1:N → Course"],
        ["contacts", "Contact[]", "No", "No", "-", "Contactos", "1:N → Contact"],
    ]
    
    for row in teacher_data:
        ws4.append(row)
    auto_adjust_columns(ws4)
    
    # HOJA 5: PROVIDER
    ws5 = wb.create_sheet("Provider")
    ws5.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws5)
    
    provider_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["name", "String", "Sí", "No", "-", "Nombre del proveedor", "-"],
        ["email", "String", "No", "No", "-", "Correo electrónico", "-"],
        ["phone", "String", "No", "No", "-", "Teléfono", "-"],
        ["address", "String", "No", "No", "-", "Dirección", "-"],
        ["taxId", "String", "No", "Sí", "-", "NIF/CIF", "-"],
        ["category", "ProviderCategory", "Sí", "No", "-", "Categoría (MATERIALS|SOFTWARE|EQUIPMENT|SERVICES|MAINTENANCE|OTHER)", "-"],
        ["description", "String", "No", "No", "-", "Descripción", "-"],
        ["website", "String", "No", "No", "-", "Sitio web", "-"],
        ["status", "ProviderStatus", "Sí", "No", "ACTIVE", "Estado (ACTIVE|INACTIVE|BLACKLISTED)", "-"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["materials", "Material[]", "No", "No", "-", "Materiales suministrados", "1:N → Material"],
        ["contacts", "Contact[]", "No", "No", "-", "Contactos", "1:N → Contact"],
    ]
    
    for row in provider_data:
        ws5.append(row)
    auto_adjust_columns(ws5)
    
    # HOJA 6: COURSE
    ws6 = wb.create_sheet("Course")
    ws6.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws6)
    
    course_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["title", "String", "Sí", "No", "-", "Título del curso", "-"],
        ["description", "String", "No", "No", "-", "Descripción", "-"],
        ["code", "String", "Sí", "Sí", "-", "Código único", "-"],
        ["level", "CourseLevel", "Sí", "No", "-", "Nivel (BEGINNER|INTERMEDIATE|ADVANCED|EXPERT)", "-"],
        ["duration", "Int", "Sí", "No", "-", "Duración en horas", "-"],
        ["maxStudents", "Int", "No", "No", "-", "Máximo de estudiantes", "-"],
        ["price", "Float", "Sí", "No", "-", "Precio del curso", "-"],
        ["isActive", "Boolean", "Sí", "No", "true", "Curso activo", "-"],
        ["startDate", "DateTime", "No", "No", "-", "Fecha de inicio", "-"],
        ["endDate", "DateTime", "No", "No", "-", "Fecha de fin", "-"],
        ["teacherId", "String", "No", "No", "-", "ID del docente", "FK → Teacher"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["teacher", "Teacher", "No", "No", "-", "Docente asignado", "N:1 ← Teacher"],
        ["enrollments", "Enrollment[]", "No", "No", "-", "Matrículas", "1:N → Enrollment"],
        ["materials", "CourseMaterial[]", "No", "No", "-", "Materiales", "1:N → CourseMaterial"],
        ["schedules", "Schedule[]", "No", "No", "-", "Horarios", "1:N → Schedule"],
        ["payments", "Payment[]", "No", "No", "-", "Pagos", "1:N → Payment"],
    ]
    
    for row in course_data:
        ws6.append(row)
    auto_adjust_columns(ws6)
    
    # HOJA 7: ENROLLMENT
    ws7 = wb.create_sheet("Enrollment")
    ws7.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws7)
    
    enrollment_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["studentId", "String", "Sí", "No", "-", "ID del estudiante", "FK → Student"],
        ["courseId", "String", "Sí", "No", "-", "ID del curso", "FK → Course"],
        ["enrollmentDate", "DateTime", "Sí", "No", "now()", "Fecha de matrícula", "-"],
        ["status", "EnrollmentStatus", "Sí", "No", "ENROLLED", "Estado (ENROLLED|IN_PROGRESS|COMPLETED|DROPPED|FAILED)", "-"],
        ["progress", "Float", "Sí", "No", "0", "Progreso (0-100)", "-"],
        ["grade", "Float", "No", "No", "-", "Calificación final", "-"],
        ["certificate", "String", "No", "No", "-", "URL del certificado", "-"],
        ["notes", "String", "No", "No", "-", "Notas adicionales", "-"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["student", "Student", "Sí", "No", "-", "Estudiante matriculado", "N:1 ← Student (Cascade)"],
        ["course", "Course", "Sí", "No", "-", "Curso matriculado", "N:1 ← Course (Cascade)"],
        ["UNIQUE", "[studentId+courseId]", "Sí", "Sí", "-", "Un estudiante solo puede matricularse una vez por curso", "-"],
    ]
    
    for row in enrollment_data:
        ws7.append(row)
    auto_adjust_columns(ws7)
    
    # HOJA 8: PAYMENT
    ws8 = wb.create_sheet("Payment")
    ws8.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws8)
    
    payment_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["studentId", "String", "No", "No", "-", "ID del estudiante", "FK → Student"],
        ["courseId", "String", "No", "No", "-", "ID del curso", "FK → Course"],
        ["amount", "Float", "Sí", "No", "-", "Monto del pago", "-"],
        ["currency", "String", "Sí", "No", "EUR", "Moneda", "-"],
        ["paymentDate", "DateTime", "Sí", "No", "now()", "Fecha del pago", "-"],
        ["paymentMethod", "PaymentMethod", "Sí", "No", "-", "Método (CASH|BANK_TRANSFER|CREDIT_CARD|DEBIT_CARD|PAYPAL|OTHER)", "-"],
        ["reference", "String", "No", "Sí", "-", "Referencia del pago", "-"],
        ["description", "String", "No", "No", "-", "Descripción", "-"],
        ["status", "PaymentStatus", "Sí", "No", "PENDING", "Estado (PENDING|PAID|OVERDUE|CANCELLED|REFUNDED)", "-"],
        ["dueDate", "DateTime", "No", "No", "-", "Fecha de vencimiento", "-"],
        ["paidDate", "DateTime", "No", "No", "-", "Fecha de pago efectivo", "-"],
        ["invoiceNumber", "String", "No", "Sí", "-", "Número de factura", "-"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["student", "Student", "No", "No", "-", "Estudiante", "N:1 ← Student (SetNull)"],
        ["course", "Course", "No", "No", "-", "Curso", "N:1 ← Course (SetNull)"],
    ]
    
    for row in payment_data:
        ws8.append(row)
    auto_adjust_columns(ws8)
    
    # HOJA 9: MATERIAL
    ws9 = wb.create_sheet("Material")
    ws9.append(["Campo", "Tipo", "Obligatorio", "Único", "Valor Defecto", "Descripción", "Relación"])
    apply_header_style(ws9)
    
    material_data = [
        ["id", "String (cuid)", "Sí", "Sí", "auto", "Identificador único", "PK"],
        ["name", "String", "Sí", "No", "-", "Nombre del material", "-"],
        ["description", "String", "No", "No", "-", "Descripción", "-"],
        ["type", "MaterialType", "Sí", "No", "-", "Tipo (BOOK|SOFTWARE|EQUIPMENT|TOOL|CONSUMABLE|DIGITAL_RESOURCE|OTHER)", "-"],
        ["quantity", "Int", "Sí", "No", "0", "Cantidad disponible", "-"],
        ["unitPrice", "Float", "No", "No", "-", "Precio unitario", "-"],
        ["location", "String", "No", "No", "-", "Ubicación física", "-"],
        ["providerId", "String", "No", "No", "-", "ID del proveedor", "FK → Provider"],
        ["isAvailable", "Boolean", "Sí", "No", "true", "Disponible", "-"],
        ["createdAt", "DateTime", "Sí", "No", "now()", "Fecha de creación", "-"],
        ["updatedAt", "DateTime", "Sí", "No", "auto", "Última actualización", "-"],
        ["provider", "Provider", "No", "No", "-", "Proveedor", "N:1 ← Provider (SetNull)"],
        ["courses", "CourseMaterial[]", "No", "No", "-", "Cursos que usan este material", "1:N → CourseMaterial"],
    ]
    
    for row in material_data:
        ws9.append(row)
    auto_adjust_columns(ws9)
    
    # HOJA 10: RELACIONES
    ws10 = wb.create_sheet("Relaciones")
    ws10.append(["Tabla Origen", "Campo", "Tipo Relación", "Tabla Destino", "Campo Destino", "Acción al Eliminar"])
    apply_header_style(ws10)
    
    relations_data = [
        ["User", "contacts", "Uno a Muchos", "Contact", "userId", "SetNull"],
        ["User", "notifications", "Uno a Muchos", "Notification", "targetUser", "SetNull"],
        ["User", "notificationPreferences", "Uno a Muchos", "NotificationPreference", "userId", "Cascade"],
        ["Student", "enrollments", "Uno a Muchos", "Enrollment", "studentId", "Cascade"],
        ["Student", "payments", "Uno a Muchos", "Payment", "studentId", "SetNull"],
        ["Student", "contacts", "Uno a Muchos", "Contact", "studentId", "SetNull"],
        ["Teacher", "courses", "Uno a Muchos", "Course", "teacherId", "SetNull"],
        ["Teacher", "contacts", "Uno a Muchos", "Contact", "teacherId", "SetNull"],
        ["Provider", "materials", "Uno a Muchos", "Material", "providerId", "SetNull"],
        ["Provider", "contacts", "Uno a Muchos", "Contact", "providerId", "SetNull"],
        ["Course", "teacher", "Muchos a Uno", "Teacher", "id", "SetNull"],
        ["Course", "enrollments", "Uno a Muchos", "Enrollment", "courseId", "Cascade"],
        ["Course", "materials", "Uno a Muchos", "CourseMaterial", "courseId", "Cascade"],
        ["Course", "schedules", "Uno a Muchos", "Schedule", "courseId", "Cascade"],
        ["Course", "payments", "Uno a Muchos", "Payment", "courseId", "SetNull"],
        ["Enrollment", "student", "Muchos a Uno", "Student", "id", "Cascade"],
        ["Enrollment", "course", "Muchos a Uno", "Course", "id", "Cascade"],
        ["Payment", "student", "Muchos a Uno", "Student", "id", "SetNull"],
        ["Payment", "course", "Muchos a Uno", "Course", "id", "SetNull"],
        ["Material", "provider", "Muchos a Uno", "Provider", "id", "SetNull"],
        ["Material", "courses", "Uno a Muchos", "CourseMaterial", "materialId", "Cascade"],
        ["CourseMaterial", "course", "Muchos a Uno", "Course", "id", "Cascade"],
        ["CourseMaterial", "material", "Muchos a Uno", "Material", "id", "Cascade"],
        ["Schedule", "course", "Muchos a Uno", "Course", "id", "Cascade"],
        ["Contact", "student", "Muchos a Uno", "Student", "id", "SetNull"],
        ["Contact", "teacher", "Muchos a Uno", "Teacher", "id", "SetNull"],
        ["Contact", "provider", "Muchos a Uno", "Provider", "id", "SetNull"],
        ["Contact", "user", "Muchos a Uno", "User", "id", "SetNull"],
        ["Notification", "user", "Muchos a Uno", "User", "id", "SetNull"],
        ["NotificationPreference", "user", "Muchos a Uno", "User", "id", "Cascade"],
    ]
    
    for row in relations_data:
        ws10.append(row)
    auto_adjust_columns(ws10)
    
    # HOJA 11: ENUMERACIONES
    ws11 = wb.create_sheet("Enumeraciones")
    ws11.append(["Enum", "Valores Posibles", "Descripción"])
    apply_header_style(ws11)
    
    enum_data = [
        ["UserRole", "ADMIN | TEACHER | STAFF", "Roles de usuario del sistema"],
        ["StudentStatus", "ACTIVE | INACTIVE | SUSPENDED | GRADUATED", "Estados de estudiantes"],
        ["ContractType", "FREELANCE | PART_TIME | FULL_TIME | HOURLY", "Tipos de contrato de docentes"],
        ["TeacherStatus", "ACTIVE | INACTIVE | ON_LEAVE", "Estados de docentes"],
        ["ProviderCategory", "MATERIALS | SOFTWARE | EQUIPMENT | SERVICES | MAINTENANCE | OTHER", "Categorías de proveedores"],
        ["ProviderStatus", "ACTIVE | INACTIVE | BLACKLISTED", "Estados de proveedores"],
        ["CourseLevel", "BEGINNER | INTERMEDIATE | ADVANCED | EXPERT", "Niveles de cursos"],
        ["EnrollmentStatus", "ENROLLED | IN_PROGRESS | COMPLETED | DROPPED | FAILED", "Estados de matrículas"],
        ["PaymentMethod", "CASH | BANK_TRANSFER | CREDIT_CARD | DEBIT_CARD | PAYPAL | OTHER", "Métodos de pago"],
        ["PaymentStatus", "PENDING | PAID | OVERDUE | CANCELLED | REFUNDED", "Estados de pagos"],
        ["MaterialType", "BOOK | SOFTWARE | EQUIPMENT | TOOL | CONSUMABLE | DIGITAL_RESOURCE | OTHER", "Tipos de materiales"],
        ["DayOfWeek", "MONDAY | TUESDAY | WEDNESDAY | THURSDAY | FRIDAY | SATURDAY | SUNDAY", "Días de la semana"],
        ["ContactCategory", "PERSONAL | WORK | EMERGENCY | ACADEMIC | ADMINISTRATIVE | TECHNICAL | OTHER", "Categorías de contactos"],
        ["SoftwareType", "LMS | VIDEO_CONFERENCE | PRODUCTIVITY | DESIGN | PROGRAMMING | ACCOUNTING | OTHER", "Tipos de software"],
        ["NotificationType", "INFO | SUCCESS | WARNING | ERROR | ALERT", "Tipos de notificaciones"],
        ["NotificationPriority", "LOW | MEDIUM | HIGH | URGENT", "Prioridades de notificaciones"],
        ["NotificationCategory", "PAYMENT | COURSE | STUDENT | TEACHER | SYSTEM | SECURITY | MAINTENANCE", "Categorías de notificaciones"],
    ]
    
    for row in enum_data:
        ws11.append(row)
    auto_adjust_columns(ws11)
    
    # Guardar archivo
    wb.save("base_datos_completa.xlsx")
    print("✅ Archivo Excel creado exitosamente: base_datos_completa.xlsx")

if __name__ == "__main__":
    create_database_excel()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Excel PLANTILLA para backup manual de datos
Sistema de Gestión Integral de Cursos
Este Excel permite rellenar datos manualmente y luego importarlos a la base de datos
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_backup_template():
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def apply_header_style(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
    
    def auto_adjust_columns(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # HOJA 0: INSTRUCCIONES
    ws_inst = wb.active
    ws_inst.title = "📖 INSTRUCCIONES"
    
    instructions = [
        ["PLANTILLA DE BACKUP - SISTEMA DE GESTIÓN DE CURSOS"],
        [""],
        ["🎯 PROPÓSITO:"],
        ["Este archivo Excel te permite mantener un backup manual de todos tus datos."],
        ["Puedes rellenar información aquí y luego importarla a la aplicación."],
        [""],
        ["📋 CÓMO USAR:"],
        ["1. Cada hoja representa una tabla de la base de datos"],
        ["2. La primera fila (azul) contiene los nombres de las columnas - NO MODIFICAR"],
        ["3. La segunda fila (amarilla) es un EJEMPLO - puedes eliminarla después"],
        ["4. A partir de la fila 3, rellena tus datos"],
        ["5. Respeta los formatos indicados en cada columna"],
        [""],
        ["⚠️ REGLAS IMPORTANTES:"],
        ["• Los campos marcados con * son OBLIGATORIOS"],
        ["• Los IDs deben ser únicos (puedes usar: EST001, PROF001, etc.)"],
        ["• Las fechas deben estar en formato: YYYY-MM-DD (ej: 2024-01-15)"],
        ["• Los campos con dropdown tienen valores predefinidos - usa SOLO esos valores"],
        ["• Los campos numéricos NO deben tener texto"],
        ["• Los emails deben ser únicos"],
        ["• Los DNI/CIF deben ser únicos"],
        [""],
        ["🔗 RELACIONES ENTRE TABLAS:"],
        ["• Cuando pongas un 'studentId' en Enrollment, debe existir en la hoja Student"],
        ["• Cuando pongas un 'courseId' en Enrollment, debe existir en la hoja Course"],
        ["• Cuando pongas un 'teacherId' en Course, debe existir en la hoja Teacher"],
        ["• Cuando pongas un 'providerId' en Material, debe existir en la hoja Provider"],
        [""],
        ["💾 IMPORTAR A LA APLICACIÓN:"],
        ["1. Guarda este archivo Excel"],
        ["2. En la aplicación, ve a 'Importar Datos'"],
        ["3. Selecciona este archivo"],
        ["4. La aplicación validará y cargará los datos"],
        [""],
        ["📊 ORDEN RECOMENDADO DE LLENADO:"],
        ["1. User (usuarios del sistema)"],
        ["2. Teacher (profesores)"],
        ["3. Student (estudiantes)"],
        ["4. Provider (proveedores)"],
        ["5. Course (cursos)"],
        ["6. Material (materiales)"],
        ["7. Enrollment (matrículas)"],
        ["8. Payment (pagos)"],
        ["9. Schedule (horarios)"],
        ["10. Contact (contactos)"],
        [""],
        ["✅ ÚLTIMA ACTUALIZACIÓN: " + datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]
    
    for i, row in enumerate(instructions, 1):
        ws_inst.append(row)
        if i == 1:
            ws_inst.row_dimensions[i].height = 25
            cell = ws_inst.cell(i, 1)
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_inst.column_dimensions['A'].width = 100
    
    # HOJA 1: USERS
    ws1 = wb.create_sheet("👤 Users")
    ws1.append(["id*", "email*", "name", "role*", "createdAt*", "NOTAS"])
    apply_header_style(ws1)
    
    # Fila de ejemplo
    ws1.append(["USR001", "admin@ejemplo.com", "Administrador Principal", "ADMIN", datetime.now().strftime("%Y-%m-%d"), "Ejemplo - puedes eliminar esta fila"])
    for cell in ws1[2]:
        cell.fill = example_fill
    
    # Validación para role
    dv_role = DataValidation(type="list", formula1='"ADMIN,TEACHER,STAFF"', allow_blank=False)
    ws1.add_data_validation(dv_role)
    dv_role.add(f"D3:D1000")
    
    auto_adjust_columns(ws1)
    
    # HOJA 2: STUDENTS
    ws2 = wb.create_sheet("🎓 Students")
    ws2.append(["id*", "name*", "email*", "phone", "address", "dni", "birthDate", "isAffiliated*", "affiliateNumber", "emergencyContact", "emergencyPhone", "medicalInfo", "status*", "createdAt*", "NOTAS"])
    apply_header_style(ws2)
    
    ws2.append([
        "EST001", 
        "Juan Pérez García", 
        "juan.perez@ejemplo.com", 
        "+34 600 123 456", 
        "Calle Mayor 123, Madrid", 
        "12345678A", 
        "1990-05-15", 
        "SI", 
        "AF001", 
        "María Pérez", 
        "+34 600 654 321", 
        "Ninguna", 
        "ACTIVE", 
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo"
    ])
    for cell in ws2[2]:
        cell.fill = example_fill
    
    # Validaciones
    dv_affiliated = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws2.add_data_validation(dv_affiliated)
    dv_affiliated.add("H3:H1000")
    
    dv_status = DataValidation(type="list", formula1='"ACTIVE,INACTIVE,SUSPENDED,GRADUATED"', allow_blank=False)
    ws2.add_data_validation(dv_status)
    dv_status.add("M3:M1000")
    
    auto_adjust_columns(ws2)
    
    # HOJA 3: TEACHERS
    ws3 = wb.create_sheet("👨‍🏫 Teachers")
    ws3.append(["id*", "name*", "email*", "phone", "address", "dni", "specialty", "experience", "cv", "contractType*", "hourlyRate", "status*", "createdAt*", "NOTAS"])
    apply_header_style(ws3)
    
    ws3.append([
        "PROF001",
        "Ana Martínez López",
        "ana.martinez@ejemplo.com",
        "+34 600 789 012",
        "Avenida Principal 45, Barcelona",
        "87654321B",
        "Programación Web",
        "10 años en desarrollo frontend",
        "https://ejemplo.com/cv-ana.pdf",
        "PART_TIME",
        "35.50",
        "ACTIVE",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo"
    ])
    for cell in ws3[2]:
        cell.fill = example_fill
    
    dv_contract = DataValidation(type="list", formula1='"FREELANCE,PART_TIME,FULL_TIME,HOURLY"', allow_blank=False)
    ws3.add_data_validation(dv_contract)
    dv_contract.add("J3:J1000")
    
    dv_teacher_status = DataValidation(type="list", formula1='"ACTIVE,INACTIVE,ON_LEAVE"', allow_blank=False)
    ws3.add_data_validation(dv_teacher_status)
    dv_teacher_status.add("L3:L1000")
    
    auto_adjust_columns(ws3)
    
    # HOJA 4: PROVIDERS
    ws4 = wb.create_sheet("🏢 Providers")
    ws4.append(["id*", "name*", "email", "phone", "address", "taxId", "category*", "description", "website", "status*", "createdAt*", "NOTAS"])
    apply_header_style(ws4)
    
    ws4.append([
        "PROV001",
        "TechBooks S.L.",
        "info@techbooks.es",
        "+34 912 345 678",
        "Polígono Industrial, Madrid",
        "B12345678",
        "MATERIALS",
        "Proveedor de libros técnicos",
        "https://techbooks.es",
        "ACTIVE",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo"
    ])
    for cell in ws4[2]:
        cell.fill = example_fill
    
    dv_category = DataValidation(type="list", formula1='"MATERIALS,SOFTWARE,EQUIPMENT,SERVICES,MAINTENANCE,OTHER"', allow_blank=False)
    ws4.add_data_validation(dv_category)
    dv_category.add("G3:G1000")
    
    dv_prov_status = DataValidation(type="list", formula1='"ACTIVE,INACTIVE,BLACKLISTED"', allow_blank=False)
    ws4.add_data_validation(dv_prov_status)
    dv_prov_status.add("J3:J1000")
    
    auto_adjust_columns(ws4)
    
    # HOJA 5: COURSES
    ws5 = wb.create_sheet("📚 Courses")
    ws5.append(["id*", "title*", "description", "code*", "level*", "duration*", "maxStudents", "price*", "isActive*", "startDate", "endDate", "teacherId", "createdAt*", "NOTAS"])
    apply_header_style(ws5)
    
    ws5.append([
        "CURSO001",
        "Desarrollo Web con React",
        "Curso completo de React desde cero",
        "DWR-2024-01",
        "INTERMEDIATE",
        "40",
        "20",
        "450.00",
        "SI",
        "2024-02-01",
        "2024-03-15",
        "PROF001",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo - teacherId debe existir en Teachers"
    ])
    for cell in ws5[2]:
        cell.fill = example_fill
    
    dv_level = DataValidation(type="list", formula1='"BEGINNER,INTERMEDIATE,ADVANCED,EXPERT"', allow_blank=False)
    ws5.add_data_validation(dv_level)
    dv_level.add("E3:E1000")
    
    dv_active = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws5.add_data_validation(dv_active)
    dv_active.add("I3:I1000")
    
    auto_adjust_columns(ws5)
    
    # HOJA 6: MATERIALS
    ws6 = wb.create_sheet("📦 Materials")
    ws6.append(["id*", "name*", "description", "type*", "quantity*", "unitPrice", "location", "providerId", "isAvailable*", "createdAt*", "NOTAS"])
    apply_header_style(ws6)
    
    ws6.append([
        "MAT001",
        "Libro JavaScript Avanzado",
        "Libro de texto para curso de JS",
        "BOOK",
        "25",
        "45.00",
        "Almacén A - Estantería 3",
        "PROV001",
        "SI",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo - providerId debe existir en Providers"
    ])
    for cell in ws6[2]:
        cell.fill = example_fill
    
    dv_mat_type = DataValidation(type="list", formula1='"BOOK,SOFTWARE,EQUIPMENT,TOOL,CONSUMABLE,DIGITAL_RESOURCE,OTHER"', allow_blank=False)
    ws6.add_data_validation(dv_mat_type)
    dv_mat_type.add("D3:D1000")
    
    dv_available = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws6.add_data_validation(dv_available)
    dv_available.add("I3:I1000")
    
    auto_adjust_columns(ws6)
    
    # HOJA 7: ENROLLMENTS
    ws7 = wb.create_sheet("📝 Enrollments")
    ws7.append(["id*", "studentId*", "courseId*", "enrollmentDate*", "status*", "progress*", "grade", "certificate", "notes", "createdAt*", "NOTAS"])
    apply_header_style(ws7)
    
    ws7.append([
        "ENROLL001",
        "EST001",
        "CURSO001",
        "2024-01-15",
        "IN_PROGRESS",
        "45.5",
        "",
        "",
        "Estudiante muy participativo",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo - IDs deben existir en Student y Course"
    ])
    for cell in ws7[2]:
        cell.fill = example_fill
    
    dv_enroll_status = DataValidation(type="list", formula1='"ENROLLED,IN_PROGRESS,COMPLETED,DROPPED,FAILED"', allow_blank=False)
    ws7.add_data_validation(dv_enroll_status)
    dv_enroll_status.add("E3:E1000")
    
    auto_adjust_columns(ws7)
    
    # HOJA 8: PAYMENTS
    ws8 = wb.create_sheet("💰 Payments")
    ws8.append(["id*", "studentId", "courseId", "amount*", "currency*", "paymentDate*", "paymentMethod*", "reference", "description", "status*", "dueDate", "paidDate", "invoiceNumber", "createdAt*", "NOTAS"])
    apply_header_style(ws8)
    
    ws8.append([
        "PAY001",
        "EST001",
        "CURSO001",
        "450.00",
        "EUR",
        "2024-01-15",
        "BANK_TRANSFER",
        "REF-2024-001",
        "Pago matrícula curso React",
        "PAID",
        "2024-01-10",
        "2024-01-15",
        "INV-2024-001",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo"
    ])
    for cell in ws8[2]:
        cell.fill = example_fill
    
    dv_payment_method = DataValidation(type="list", formula1='"CASH,BANK_TRANSFER,CREDIT_CARD,DEBIT_CARD,PAYPAL,OTHER"', allow_blank=False)
    ws8.add_data_validation(dv_payment_method)
    dv_payment_method.add("G3:G1000")
    
    dv_payment_status = DataValidation(type="list", formula1='"PENDING,PAID,OVERDUE,CANCELLED,REFUNDED"', allow_blank=False)
    ws8.add_data_validation(dv_payment_status)
    dv_payment_status.add("J3:J1000")
    
    auto_adjust_columns(ws8)
    
    # HOJA 9: SCHEDULES
    ws9 = wb.create_sheet("🕐 Schedules")
    ws9.append(["id*", "courseId*", "dayOfWeek*", "startTime*", "endTime*", "classroom", "isRecurring*", "notes", "createdAt*", "NOTAS"])
    apply_header_style(ws9)
    
    ws9.append([
        "SCH001",
        "CURSO001",
        "MONDAY",
        "2024-01-15 09:00:00",
        "2024-01-15 11:00:00",
        "Aula 101",
        "SI",
        "Clase teórica",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo - courseId debe existir en Courses"
    ])
    for cell in ws9[2]:
        cell.fill = example_fill
    
    dv_day = DataValidation(type="list", formula1='"MONDAY,TUESDAY,WEDNESDAY,THURSDAY,FRIDAY,SATURDAY,SUNDAY"', allow_blank=False)
    ws9.add_data_validation(dv_day)
    dv_day.add("C3:C1000")
    
    dv_recurring = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws9.add_data_validation(dv_recurring)
    dv_recurring.add("G3:G1000")
    
    auto_adjust_columns(ws9)
    
    # HOJA 10: CONTACTS
    ws10 = wb.create_sheet("📞 Contacts")
    ws10.append(["id*", "name*", "email", "phone", "mobile", "address", "company", "position", "category*", "notes", "isPrimary*", "studentId", "teacherId", "providerId", "userId", "createdAt*", "NOTAS"])
    apply_header_style(ws10)
    
    ws10.append([
        "CONT001",
        "María Pérez (Madre)",
        "maria.perez@ejemplo.com",
        "+34 600 111 222",
        "+34 600 111 222",
        "Calle Mayor 123",
        "",
        "",
        "EMERGENCY",
        "Contacto de emergencia de Juan",
        "SI",
        "EST001",
        "",
        "",
        "",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo - Solo rellenar UNO de: studentId, teacherId, providerId o userId"
    ])
    for cell in ws10[2]:
        cell.fill = example_fill
    
    dv_contact_cat = DataValidation(type="list", formula1='"PERSONAL,WORK,EMERGENCY,ACADEMIC,ADMINISTRATIVE,TECHNICAL,OTHER"', allow_blank=False)
    ws10.add_data_validation(dv_contact_cat)
    dv_contact_cat.add("I3:I1000")
    
    dv_primary = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws10.add_data_validation(dv_primary)
    dv_primary.add("K3:K1000")
    
    auto_adjust_columns(ws10)
    
    # HOJA 11: SOFTWARE
    ws11 = wb.create_sheet("💻 Software")
    ws11.append(["id*", "name*", "version", "type*", "license", "licenseKey", "expiryDate", "provider", "description", "isActive*", "maxUsers", "currentUsers*", "url", "createdAt*", "NOTAS"])
    apply_header_style(ws11)
    
    ws11.append([
        "SOFT001",
        "Zoom Education",
        "5.14.0",
        "VIDEO_CONFERENCE",
        "Educativa Anual",
        "ZOOM-EDU-2024-XXXXX",
        "2024-12-31",
        "Zoom Video Communications",
        "Plataforma de videoconferencias",
        "SI",
        "100",
        "45",
        "https://zoom.us",
        datetime.now().strftime("%Y-%m-%d"),
        "Ejemplo"
    ])
    for cell in ws11[2]:
        cell.fill = example_fill
    
    dv_soft_type = DataValidation(type="list", formula1='"LMS,VIDEO_CONFERENCE,PRODUCTIVITY,DESIGN,PROGRAMMING,ACCOUNTING,OTHER"', allow_blank=False)
    ws11.add_data_validation(dv_soft_type)
    dv_soft_type.add("D3:D1000")
    
    dv_soft_active = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws11.add_data_validation(dv_soft_active)
    dv_soft_active.add("J3:J1000")
    
    auto_adjust_columns(ws11)
    
    # Guardar archivo
    filename = "PLANTILLA_BACKUP_DATOS.xlsx"
    wb.save(filename)
    print(f"✅ Plantilla de backup creada exitosamente: {filename}")
    print(f"📊 Hojas creadas: {len(wb.sheetnames)}")
    print(f"📝 Puedes empezar a rellenar datos a partir de la fila 3 de cada hoja")

if __name__ == "__main__":
    create_backup_template()
